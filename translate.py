"""
    This script was created to facilitate the translation of:
        Baldur's Gate II Enhanced Edition.
"""

import os
import json
import re
import xlsxwriter
import openpyxl
from xml.etree import ElementTree

# Current version of script is optimized to source from english language
SOURCE_LOCALE = 'en'

# Change to you target language
TARGET_LOCALE = 'pt'
TARGET_NAME = 'portuguese'
TARGET_VERBOSE = 'Português (Brasil)'

REGEX_PLAYER = re.compile(r"\[\s*Player\s+Name\s*\]")
REGEX_URL_TAG = re.compile(r"\[/*url['\w\s=:]*\]")
REGEX_FORMATABLE = re.compile(r"{\d+}")
REGEX_SPECIAL_CARACTERES = re.compile('[\n\r"]+')

XML_EXTENSION = '.stringtable'
MAX_XLSX_LINES = 35000
JSON_FILE = 'translate_data.json'

DATA = (
    ('data', 'exported', 'localized'),
    ('data_vx1', 'exported', 'localized'),
    ('data_vx2', 'exported', 'localized'),
    ('data_vx3', 'exported', 'localized'),
)

TEMP_FOLDER = 'temp'
LANGUAGE_FILE = 'language.xml'


def title_is_valid(chunk, title):
    if title[:2] == "I'":
        return False
    c, l = chunk.strip("'").strip('"').strip(), len(title)
    if l <= 3:
        return False
    words = re.findall(r'[A-Z][a-z]+', title)
    if c[:l] == title and len(c[l:]) and len(words) == 1:
        return False
    return True


def get_titles_words(text):
    titles = []
    for chunk in re.split(r'[\[\],\.\n]', text):
        title = None
        lower = None
        for word in chunk.strip().split(' '):
            match = re.match(r"[A-Z][a-z]*'?[a-z]", word.strip("'").strip('"'))
            match_l = re.match(r"^[a-z]{1,3}(?!\w)", word.strip("'").strip('"'))
            if match:
                title = (
                    f'{title} {lower} {match.group()}'
                    if (title and lower) else (
                        f'{title} {match.group()}'
                        if title else match.group()
                    )
                )
                lower = None
            elif match_l:
                if lower:
                    title = None
                    lower = None
                else:
                    lower = match_l.group()
            elif title:
                if title_is_valid(chunk, title):
                    titles.append(title)
                title = None
                lower = None
        if title:
            if title_is_valid(chunk, title):
                titles.append(title)
    return titles


def set_immutable(text):
    if not text:
        return (None, None, None)

    immutable = set(
        REGEX_PLAYER.findall(text) +
        REGEX_URL_TAG.findall(text) +
        REGEX_FORMATABLE.findall(text) +
        REGEX_SPECIAL_CARACTERES.findall(text)
    )

    immutable_keys = {}
    for index, value in enumerate(immutable):
        key = str(index).rjust(4, '0')
        immutable_keys[key] = value
        text = text.replace(value, f"[{key}]")

    titles = {}
    for index, title in enumerate(set(get_titles_words(text))):
        key = f"00-{str(index).rjust(2, '0')}"
        titles[key] = title
        text = text.replace(title, f"[{key}]")

    return (text, immutable_keys, titles)


def revert_immutable(text, immutable_keys, titles):
    for key, value in immutable_keys.items():
        text = re.sub(f'\[\s*{key}\s*\]', value, text)
    text = text.strip()
    if text[:2] == '" ':
        text = f'"{text[2:]}'
    for k, v in titles.items():
        text = re.sub(f'\[\s*00\s*-\s*{k.split("-")[-1]}\s*\]', v, text)
    return text


def get_files(folder, extension):
    files = []
    for f in os.listdir(folder):
        path = os.path.join(folder, f)
        if os.path.isdir(path):
            files.extend(get_files(path, extension))
        if os.path.isfile(path) and path[-len(extension):] == extension:
            files.append(path)
    return files


def generate_xlsx(data, translate_data, locale):
    part_num = 1
    current_line = 1
    current_title = 1

    xml_base = f'{data[0]}_translate'
    path = os.path.join(*data, locale)
    translate_data[path] = {}

    title_wb = xlsxwriter.Workbook(
        os.path.join(TEMP_FOLDER, f'{xml_base}.titles.xlsx')
    )
    title_ws = title_wb.add_worksheet()
    title_writes = {}
    workbook = xlsxwriter.Workbook(
        os.path.join(TEMP_FOLDER, f'{xml_base}.part{part_num}.xlsx')
    )
    worksheet = workbook.add_worksheet()

    for file in get_files(path, XML_EXTENSION):
        file_key = file.replace(path, '').strip(os.path.sep)
        translate_data[path][file_key] = {}
        xml = ElementTree.parse(file)
        for entry in xml.getroot().find('Entries').findall('Entry'):
            id = entry.find('ID').text
            for text_key in ['DefaultText', 'FemaleText']:
                line_key = f'A{current_line}'
                text, immutable, titles = set_immutable(
                    entry.find(text_key).text
                )

                if text is None:
                    continue

                worksheet.write(line_key, text)

                titles_lines = {}
                for key, title in titles.items():
                    t_l_key = title_writes.get(title, None)
                    if not t_l_key:
                        t_l_key = f'A{current_title}'
                        title_writes[title] = t_l_key
                        title_ws.write(t_l_key, title)
                        current_title += 1
                    titles_lines[key] = t_l_key

                translate_data[path][file_key][id] = (
                    translate_data[path][file_key].get(id, {})
                )
                translate_data[path][file_key][id].update({
                    text_key: {
                        'xlsx': f'{xml_base}.part{part_num}.xlsx',
                        'xlsx_titles': f'{xml_base}.titles.xlsx',
                        'line': line_key,
                        'immutable': immutable,
                        'titles': titles_lines,
                    }
                })

                current_line += 1
                if current_line > MAX_XLSX_LINES:
                    part_num += 1
                    current_line = 1
                    workbook.close()
                    print(f'Created: {workbook.filename}')
                    workbook = xlsxwriter.Workbook(
                        os.path.join(
                            TEMP_FOLDER, f'{xml_base}.part{part_num}.xlsx')
                    )
                    worksheet = workbook.add_worksheet()
    workbook.close()
    print(f'Created: {workbook.filename}')
    title_wb.close()
    print(f'Created: {title_wb.filename}')


def generate_locale(data, translate_data, source, target, name, verbose):
    language = ElementTree.parse(
        os.path.join(*data[0], source, LANGUAGE_FILE)
    )
    lang_child = language.getroot()
    lang_child.find('Name').text = TARGET_NAME
    lang_child.find('GUIString').text = TARGET_VERBOSE
    os.makedirs(os.path.join(*data[0], target), exist_ok=True)
    language.write(
        os.path.join(*data[0], target, LANGUAGE_FILE),
        encoding="utf-8", xml_declaration=True
    )

    xlsx_files = {}

    for path, files in translate_data.items():
        for file, entries in files.items():
            xml = ElementTree.parse(
                os.path.join(path, file)
            )
            root = xml.getroot()
            for entry in root.find('Entries').findall('Entry'):
                id = entry.find('ID').text
                for text_key in ['DefaultText', 'FemaleText']:
                    data = entries.get(id, {}).get(text_key, None)
                    if data is None:
                        continue

                    worksheet = xlsx_files.get(data['xlsx'], None)
                    if not worksheet:
                        workbook = openpyxl.load_workbook(
                            os.path.join(TEMP_FOLDER, data['xlsx'])
                        )
                        worksheet = workbook.active
                        xlsx_files[data['xlsx']] = worksheet

                    title_ws = xlsx_files.get(data['xlsx_titles'], None)
                    if not title_ws:
                        title_wb = openpyxl.load_workbook(
                            os.path.join(TEMP_FOLDER, data['xlsx_titles'])
                        )
                        title_ws = title_wb.active
                        xlsx_files[data['xlsx_titles']] = title_ws

                    titles = {}
                    for key, line in data['titles'].items():
                        titles[key] = (title_ws[line].value or '').strip()

                    entry.find(text_key).text = revert_immutable(
                        (worksheet[data['line']].value or ''),
                        data['immutable'],
                        titles,
                    )

            file_output = os.path.join(
                path.strip(source).strip(os.path.sep), target, file
            )
            os.makedirs(
                os.path.join(*file_output.split(os.path.sep)[:-1]),
                exist_ok=True
            )
            xml.write(file_output, encoding="utf-8", xml_declaration=True)
            print('Created:', file_output)


if __name__ == '__main__':
    json_path = os.path.join(TEMP_FOLDER, JSON_FILE)
    if not os.path.isfile(json_path):
        translate_data = {}

        print('Generating files:')
        for data in DATA:
            generate_xlsx(data, translate_data, SOURCE_LOCALE)

        with open(json_path, 'wb') as file:
            file.write(json.dumps(translate_data).encode('utf8'))
            print(f'Created:', JSON_FILE)

        print(
            'Translate XLSX files using https://translate.google.com.br/\n'
            f'Replace new XLSX files in {TEMP_FOLDER} '
            'folder before re-running this script'
        )
    else:
        with open(json_path, 'rb') as file:
            translate_data = json.loads(file.read().decode('utf8'))

        print(f'Translate to "{TARGET_VERBOSE}" in progress:')
        generate_locale(
            DATA, translate_data, SOURCE_LOCALE,
            TARGET_LOCALE, TARGET_NAME, TARGET_VERBOSE)
        print('Translate finish.')
